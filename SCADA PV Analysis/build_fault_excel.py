"""
Build fault_knowledge_base.xlsx from fault_knowledge_base.json
Run:  python build_fault_excel.py
"""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\RichardMUSI\OneDrive - Dolfines\Bureau\PVPAT")
SRC  = BASE / "fault_knowledge_base.json"
DST  = BASE / "fault_knowledge_base.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "navy":        "1F3864",
    "navy_light":  "2F5496",
    "orange":      "E67E22",
    "red":         "C0392B",
    "amber":       "F39C12",
    "green":       "1A6B35",
    "grey_hdr":    "D6DCE4",
    "grey_row":    "F2F2F2",
    "white":       "FFFFFF",
    "HIGH":        "FDECEA",
    "HIGH_txt":    "B71C1C",
    "MEDIUM":      "FFF3E0",
    "MEDIUM_txt":  "BF6000",
    "LOW":         "E8F5EC",
    "LOW_txt":     "1A6B35",
    "INFO":        "EEF3FA",
    "INFO_txt":    "1F4E79",
}

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, color="000000", size=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

thin = Side(style="thin", color="BFBFBF")
thick = Side(style="medium", color="808080")

def border_thin():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def border_thick_bottom():
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def write_header_row(ws, row_idx, headers, widths, bg="navy"):
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row_idx, column=ci, value=h)
        cell.fill    = fill(C[bg])
        cell.font    = font(bold=True, color=C["white"], size=9)
        cell.alignment = align(wrap=True, h="center", v="center")
        cell.border  = border_thick_bottom()
        set_col_width(ws, ci, w)

def style_data_cell(cell, sev=None, bold=False, italic=False, h="left"):
    bg  = C.get(sev, C["white"])
    txt = C.get(f"{sev}_txt", "000000") if sev else "000000"
    cell.fill      = fill(bg)
    cell.font      = font(bold=bold, color=txt, size=8.5, italic=italic)
    cell.alignment = align(wrap=True, h=h, v="top")
    cell.border    = border_thin()

def list_to_str(v):
    if isinstance(v, list):
        return "\n• ".join(["• " + x.strip() if i == 0 else x.strip()
                             for i, x in enumerate(v)])
    return str(v) if v is not None else ""


# ── Load JSON ─────────────────────────────────────────────────────────────────
with open(SRC, encoding="utf-8") as f:
    kb = json.load(f)

wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default blank sheet


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — INVERTER FAULTS (one row per fault, all manufacturers)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Inverter Faults")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

headers = ["ID", "Manufacturer", "Models", "Fault Code", "Fault Name",
           "Category", "Severity", "Frequency",
           "Symptoms", "Root Causes", "Diagnostic Actions",
           "MTTF Benchmark (days)", "Notes / Platform"]
widths  = [9, 14, 28, 16, 28, 14, 10, 16, 45, 45, 45, 14, 38]

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
title = ws.cell(row=1, column=1,
                value="SOLAR PV — INVERTER FAULT KNOWLEDGE BASE")
title.fill      = fill(C["navy"])
title.font      = font(bold=True, color=C["white"], size=11)
title.alignment = align(h="center", v="center", wrap=False)
ws.row_dimensions[1].height = 22

write_header_row(ws, 2, headers, widths)
ws.row_dimensions[2].height = 30

row = 3
for mfr in kb["inverter_manufacturers"]:
    models_str = ", ".join(mfr.get("common_models", []))
    platform   = mfr.get("monitoring_platform", "")
    for fault in mfr.get("faults", []):
        sev = fault.get("severity", "INFO")
        data = [
            fault.get("id", ""),
            mfr["manufacturer"],
            models_str,
            fault.get("code", ""),
            fault.get("name", ""),
            fault.get("category", ""),
            sev,
            fault.get("frequency", ""),
            list_to_str(fault.get("symptoms", [])),
            list_to_str(fault.get("root_causes", [])),
            list_to_str(fault.get("actions", [])),
            fault.get("mttf_benchmark_days", ""),
            f"{platform}\n{fault.get('note', '')}".strip(),
        ]
        for ci, val in enumerate(data, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            # Severity column gets colour badge; rest get row-zebra
            if ci == 7:
                style_data_cell(cell, sev=sev, bold=True, h="center")
            else:
                row_bg = C["grey_row"] if row % 2 == 0 else C["white"]
                cell.fill      = fill(row_bg)
                cell.font      = font(size=8.5)
                cell.alignment = align(wrap=True, h="left", v="top")
                cell.border    = border_thin()
        ws.row_dimensions[row].height = 72
        row += 1

ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — MODULE ISSUES (one row per issue, all technologies)
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Module Issues")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

headers2 = ["ID", "Technology", "Manufacturers", "Issue Name",
            "Severity", "Frequency",
            "Symptoms", "Root Causes", "Corrective Actions",
            "Energy Impact", "Warranty / Notes"]
widths2  = [12, 22, 38, 30, 10, 18, 42, 42, 42, 30, 38]

ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers2))
title2 = ws2.cell(row=1, column=1,
                  value="SOLAR PV — MODULE DEGRADATION & FAULT KNOWLEDGE BASE")
title2.fill      = fill(C["navy"])
title2.font      = font(bold=True, color=C["white"], size=11)
title2.alignment = align(h="center", v="center", wrap=False)
ws2.row_dimensions[1].height = 22

write_header_row(ws2, 2, headers2, widths2)
ws2.row_dimensions[2].height = 30

row2 = 3
for tech in kb["module_technologies"]:
    tech_name  = tech["technology"]
    mfrs_str   = ", ".join(tech.get("manufacturers", []))
    pid_note   = f"PID risk: {tech.get('pid_risk','—')}"
    lid_note   = f"LID: {tech.get('lid_effect','—')}"
    deg_note   = f"Degradation: {tech.get('typical_degradation_pct_per_year','—')}%/yr"
    tech_notes = f"{pid_note} | {lid_note} | {deg_note}"

    for issue in tech.get("issues", []):
        sev = issue.get("severity", "INFO")
        wref = issue.get("warranty_reference", "")
        note_col = "\n".join(filter(None, [tech_notes, wref]))
        data2 = [
            issue.get("id", ""),
            tech_name,
            mfrs_str,
            issue.get("name", ""),
            sev,
            issue.get("frequency", ""),
            list_to_str(issue.get("symptoms", [])),
            list_to_str(issue.get("root_causes", [])),
            list_to_str(issue.get("actions", [])),
            issue.get("energy_impact", ""),
            note_col,
        ]
        for ci, val in enumerate(data2, start=1):
            cell = ws2.cell(row=row2, column=ci, value=val)
            if ci == 5:
                style_data_cell(cell, sev=sev, bold=True, h="center")
            else:
                row_bg = C["grey_row"] if row2 % 2 == 0 else C["white"]
                cell.fill      = fill(row_bg)
                cell.font      = font(size=8.5)
                cell.alignment = align(wrap=True, h="left", v="top")
                cell.border    = border_thin()
        ws2.row_dimensions[row2].height = 72
        row2 += 1

ws2.auto_filter.ref = f"A2:{get_column_letter(len(headers2))}2"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — DIAGNOSTIC MATRIX
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Diagnostic Matrix")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

headers3 = ["SCADA Symptom Observed",
            "Investigate First (Fault IDs / Categories)",
            "Tool / Method Required"]
widths3  = [60, 65, 45]

ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers3))
title3 = ws3.cell(row=1, column=1,
                  value="SCADA SYMPTOM → DIAGNOSTIC MATRIX (All Brands / Technologies)")
title3.fill      = fill(C["navy"])
title3.font      = font(bold=True, color=C["white"], size=11)
title3.alignment = align(h="center", v="center", wrap=False)
ws3.row_dimensions[1].height = 22

write_header_row(ws3, 2, headers3, widths3)
ws3.row_dimensions[2].height = 28

row3 = 3
for i, entry in enumerate(kb["diagnostic_matrix"]["entries"]):
    investigate = list_to_str(entry.get("investigate_first",
                              entry.get("first_check", [])))
    data3 = [
        entry.get("scada_symptom", ""),
        investigate,
        entry.get("tool_required", ""),
    ]
    row_bg = C["grey_row"] if row3 % 2 == 0 else C["white"]
    for ci, val in enumerate(data3, start=1):
        cell = ws3.cell(row=row3, column=ci, value=val)
        cell.fill      = fill(row_bg)
        cell.font      = font(size=9, bold=(ci == 1))
        cell.alignment = align(wrap=True, h="left", v="top")
        cell.border    = border_thin()
    ws3.row_dimensions[row3].height = 60
    row3 += 1


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — O&M SCHEDULE
# ═════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("O&M Schedule")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A3"

headers4 = ["Frequency", "Task Description", "Tools / Resources Required"]
widths4  = [16, 75, 40]

ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers4))
title4 = ws4.cell(row=1, column=1,
                  value="RECOMMENDED O&M INSPECTION SCHEDULE — UTILITY-SCALE PV")
title4.fill      = fill(C["navy"])
title4.font      = font(bold=True, color=C["white"], size=11)
title4.alignment = align(h="center", v="center", wrap=False)
ws4.row_dimensions[1].height = 22

write_header_row(ws4, 2, headers4, widths4)
ws4.row_dimensions[2].height = 28

# Colour-code by frequency
freq_colours = {
    "Daily":        ("E8F4FD", "1A5276"),
    "Monthly":      ("E8F8F5", "1A6B35"),
    "Quarterly":    ("FEF9E7", "7D6608"),
    "Biannual":     ("FDF2E9", "935116"),
    "Annual":       ("FDEDEC", "922B21"),
    "Every 2 years":("F5EEF8", "6C3483"),
    "Every 5 years":("F2F3F4", "424949"),
}

row4 = 3
for task in kb["o_and_m_schedule"]["tasks"]:
    freq = task["frequency"]
    bg, txt_col = freq_colours.get(freq, (C["white"], "000000"))
    data4 = [freq, task["task"], task["tools"]]
    for ci, val in enumerate(data4, start=1):
        cell = ws4.cell(row=row4, column=ci, value=val)
        cell.fill      = fill(bg)
        cell.font      = font(size=9, bold=(ci == 1), color=txt_col)
        cell.alignment = align(wrap=True, h="left", v="top")
        cell.border    = border_thin()
    ws4.row_dimensions[row4].height = 45
    row4 += 1


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 — EXPECTED SEASONAL PATTERNS
# ═════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Expected Patterns")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "A3"

headers5 = ["ID", "Pattern Name", "Category",
            "Explanation", "Correct Interpretation",
            "KPI Impact", "Diagnostic Test",
            "Applicable Technologies",
            "Flag If (Abnormal)", "Do NOT Flag If (Normal)"]
widths5  = [10, 32, 20, 55, 50, 40, 50, 35, 45, 45]

ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers5))
title5 = ws5.cell(row=1, column=1,
                  value="EXPECTED SEASONAL PATTERNS — Normal Behaviours (Do Not Flag as Faults)")
title5.fill      = fill(C["green"])
title5.font      = font(bold=True, color=C["white"], size=11)
title5.alignment = align(h="center", v="center", wrap=False)
ws5.row_dimensions[1].height = 22

write_header_row(ws5, 2, headers5, widths5, bg="navy")
ws5.row_dimensions[2].height = 30

cat_colours = {
    "Normal — seasonal":               ("E8F8F5", "1A6B35"),
    "Expected — seasonal O&M":         ("FEF9E7", "7D6608"),
    "Normal — beneficial seasonal effect": ("E8F4FD", "1A5276"),
    "Normal — MPPT / irradiance threshold": ("F5EEF8", "6C3483"),
    "Expected — soiling washoff":      ("FDF2E9", "935116"),
    "Expected — thermal derating":     ("FDEDEC", "922B21"),
}

row5 = 3
for i, pat in enumerate(kb["expected_seasonal_patterns"]["patterns"]):
    cat = pat.get("category", "")
    bg, txt_col = cat_colours.get(cat, (C["white"], "000000"))
    data5 = [
        pat.get("id", ""),
        pat.get("pattern", ""),
        cat,
        pat.get("explanation", ""),
        pat.get("correct_interpretation", ""),
        pat.get("kpi_impact", ""),
        pat.get("diagnostic_test", ""),
        list_to_str(pat.get("applicable_technologies", [])),
        pat.get("flag_if", ""),
        pat.get("do_not_flag_if", ""),
    ]
    for ci, val in enumerate(data5, start=1):
        cell = ws5.cell(row=row5, column=ci, value=val)
        cell.fill      = fill(bg)
        cell.font      = font(size=8.5, bold=(ci in (1, 2)), color=txt_col if ci <= 3 else "000000")
        cell.alignment = align(wrap=True, h="left", v="top")
        cell.border    = border_thin()
    ws5.row_dimensions[row5].height = 80
    row5 += 1

ws5.auto_filter.ref = f"A2:{get_column_letter(len(headers5))}2"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 — SUMMARY / COVER
# ═════════════════════════════════════════════════════════════════════════════
ws0 = wb.create_sheet("README", 0)   # insert at position 0
ws0.sheet_view.showGridLines = False

def cov(row, col, val, bg="navy", txt="white", sz=10, bold=False,
        italic=False, h="left", merge_to=None):
    cell = ws0.cell(row=row, column=col, value=val)
    cell.fill      = fill(C.get(bg, bg))
    cell.font      = font(bold=bold, color=C.get(txt, txt), size=sz, italic=italic)
    cell.alignment = align(h=h, v="center", wrap=True)
    if merge_to:
        ws0.merge_cells(start_row=row, start_column=col,
                        end_row=row, end_column=merge_to)
    return cell

for ci in range(1, 4):
    ws0.column_dimensions[get_column_letter(ci)].width = [4, 38, 55][ci-1]

ws0.row_dimensions[1].height = 8
cov(2, 1, "SOLAR PV FAULT & DEGRADATION KNOWLEDGE BASE",
    bg="navy", txt="white", sz=14, bold=True, h="center", merge_to=3)
ws0.row_dimensions[2].height = 32

cov(3, 1, "Multi-Manufacturer Inverter & Module Reference  |  Version 1.1  |  2026",
    bg="navy_light", txt="white", sz=9, italic=True, h="center", merge_to=3)
ws0.row_dimensions[3].height = 18

ws0.row_dimensions[4].height = 10

cov(5, 2, "CONTENTS", bg="grey_hdr", txt="navy", sz=10, bold=True, merge_to=3)
ws0.row_dimensions[5].height = 20

sheets_desc = [
    ("Inverter Faults",    "One row per fault across 6 inverter brands: Sungrow, SMA, Huawei, ABB/FIMER, Fronius, Ingeteam"),
    ("Module Issues",      "One row per failure mode across 4 module technologies: CdTe, Mono-Si PERC/TOPCon/HJT, IBC, Bifacial"),
    ("Diagnostic Matrix",  "SCADA symptom → fault hypothesis lookup table — quick triage reference for O&M teams"),
    ("O&M Schedule",       "Recommended inspection frequency and task checklist for utility-scale PV"),
    ("Expected Patterns",  "Normal seasonal behaviours that should NOT be flagged as underperformance — avoid false alarms"),
]
for i, (sheet, desc) in enumerate(sheets_desc, start=6):
    cov(i, 2, f"  {sheet}", bg="white", txt="navy_light", sz=9, bold=True)
    cov(i, 3, desc, bg="white", txt="000000", sz=9)
    ws0.row_dimensions[i].height = 18

ws0.row_dimensions[10] = ws0.row_dimensions.get(10, openpyxl.worksheet.dimensions.RowDimension(ws0, index=10))
ws0.row_dimensions[10].height = 12

cov(11, 2, "HOW TO USE", bg="grey_hdr", txt="navy", sz=10, bold=True, merge_to=3)
ws0.row_dimensions[11].height = 20

usage = [
    "1. Use 'Inverter Faults' or 'Module Issues' tabs to look up a known fault code or failure mode.",
    "2. Use 'Diagnostic Matrix' tab when you have a SCADA symptom but don't know the fault — maps observation to investigation path.",
    "3. Filter by Severity (HIGH / MEDIUM / LOW / INFO) to prioritise corrective actions.",
    "4. All action items are practical field steps — cross-reference with OEM service manual for specific firmware versions.",
    "5. O&M Schedule tab provides a recommended inspection calendar to prevent faults before they occur.",
]
for i, txt in enumerate(usage, start=12):
    cov(i, 2, txt, bg="white", txt="000000", sz=8.5, merge_to=3)
    ws0.row_dimensions[i].height = 16

row_note = 12 + len(usage) + 1
cov(row_note, 2,
    "Note: Fault codes and thresholds are indicative — always verify against the specific firmware version and "
    "national grid code applicable to the installation.",
    bg="grey_row", txt="000000", sz=8, italic=True, merge_to=3)
ws0.row_dimensions[row_note].height = 28


# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(DST)
print(f"Saved: {DST}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
inverter_rows  = sum(len(m['faults']) for m in kb['inverter_manufacturers'])
module_rows    = sum(len(t['issues']) for t in kb['module_technologies'])
diag_rows      = len(kb['diagnostic_matrix']['entries'])
pattern_rows   = len(kb['expected_seasonal_patterns']['patterns'])
print(f"  Inverter fault rows   : {inverter_rows}")
print(f"  Module issue rows     : {module_rows}")
print(f"  Diagnostic rows       : {diag_rows}")
print(f"  Seasonal pattern rows : {pattern_rows}")
